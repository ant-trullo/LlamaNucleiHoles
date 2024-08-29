"""This function loads previously done analysis.

Inoput are the analysis folder path and the raw data.
All the output are the variables of the software (matrix video, analysis parameters values etc..)
"""


import numpy as np
from openpyxl import load_workbook
from scipy.signal import argrelextrema
from skimage.filters import gaussian

import SaveReadMatrix
import MultiLoadCzi5D
import UsefulWidgets


class RawDataLoader:
    """This class loads raw data and shapes them accordingly to the analysis done."""
    def __init__(self, analysis_folder, fnames):

        raw_data     =  MultiLoadCzi5D.MultiLoadCzi5D(fnames, np.load(analysis_folder + '/nucs_spots_channels.npy'))
        im_red_smpl  =  np.load(analysis_folder + '/im_red_smpl.npy')
        jj_start     =  np.where(np.sum(raw_data.imarray_red - im_red_smpl[0], axis=(1, 2)) == 0)[0][0]
        jj_end       =  np.where(np.sum(raw_data.imarray_red - im_red_smpl[1], axis=(1, 2)) == 0)[0][0]

        self.imarray_green    =  raw_data.imarray_green[jj_start:jj_end + 1]
        self.imarray_red      =  raw_data.imarray_red[jj_start:jj_end + 1]
        self.green4d          =  raw_data.green4d[jj_start:jj_end + 1]
        self.red4d            =  raw_data.red4d[jj_start:jj_end + 1]
        self.pix_size_x       =  raw_data.pix_size_x
        self.pix_size_z       =  raw_data.pix_size_z
        self.time_step_value  =  raw_data.time_step_value
        self.fnames           =  fnames


class SpotsIntsVol:
    """Loads intensity and volume of detected spots."""
    def __init__(self, analysis_folder):

        self.spots_ints    =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_3d_ints.npy').spts_lbls
        self.spots_vol     =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_3d_vol.npy').spts_lbls
        self.spots_tzxy    =  np.load(analysis_folder + '/spots_3d_tzxy.npy')
        self.spots_coords  =  np.load(analysis_folder + '/spots_3d_coords.npy')


class AnalysisParameters:
    """Load analysis parameters."""
    def __init__(self, analysis_folder):

        wb       =  load_workbook(analysis_folder + '/journal.xlsx')
        s_wb     =  wb["Info"]

        self.gauss_kernelsize_value  =  s_wb.cell(1, 2).value
        self.spots_thr_value         =  s_wb.cell(2, 2).value
        self.volume_thr_value        =  s_wb.cell(3, 2).value
        self.dist_thr_value          =  s_wb.cell(4, 2).value
        self.max_dist                =  s_wb.cell(5, 2).value


class NucleiSegmented:
    """Load nuclei segmented results."""
    def __init__(self, analysis_folder, green4d):

        nucs_lbld  =  np.load(analysis_folder + '/nuclei_segmented.npy')

        pbar  =  UsefulWidgets.ProgressBar(total1=nucs_lbld.shape[0])
        pbar.show()
        pbar.update_progressbar(0)

        green_minp  =  np.zeros(nucs_lbld.shape, dtype=green4d.dtype)                                    # initialize the image-matrix for the intensity projection (sum in z)
        mxxs        =  []
        for tt in range(nucs_lbld.shape[0]):                                                                                          # for each time step
            pbar.update_progressbar(tt + 1)
            z_prof          =  np.sum(green4d[tt], axis=(1, 2))
            mxx             =  argrelextrema(z_prof, np.greater)[0]
            if mxx.size == 2:
                mxxs.append(mxx)
                green_bff       =  gaussian(green4d[tt, mxx[0]:mxx[1]].astype(np.float32), 1.5)                                            # gaussian smoothing
                green_minp[tt]  =  green_bff.sum(0)                                                                         # sum over z

        mxxs      =  np.asarray(mxxs)
        z_ref     =  mxxs.mean(0)
        z_ref[0]  =  np.round(z_ref[0])
        z_ref[1]  =  np.round(z_ref[1])

        self.nucs_lbld   =  nucs_lbld
        self.green_minp  =  green_minp
        self.z_ref       =  z_ref.astype(int)
