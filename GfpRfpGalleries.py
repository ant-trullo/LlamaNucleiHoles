"""This function rapresents the behaviour of GFP around RFP spots.

Input is the analysis folder with CheckIntensityAroundSpots results.
Output is gallery.
"""

import numpy as np
import pyqtgraph as pg
import time
from openpyxl import load_workbook


class GfpRfpGalleries:
    """This function reads xlsx files and plots results."""
    def __init__(self, analysis_folder):

        spts_book   =  load_workbook(analysis_folder + '/journal.xlsx')
        # spts_sheet  =  spts_book.get_sheet_by_name('Ints')
        spts_sheet  =  spts_book.get_sheet_by_name('Ints by Background')

        gfp_book   =  load_workbook(analysis_folder + '/GfpIntensityAroundSpots.xlsx')
        gfp_sheet  =  gfp_book.get_sheet_by_name('Gfp Intensity Around Spots')

        tags  =  []
        for pp in range(1, gfp_sheet.max_column):
            tags.append(gfp_sheet.cell(row=1, column=pp + 1).value[4:])
        tags  =  np.asarray(tags)

        spts_gfp_mtx  =  np.zeros((2, gfp_sheet.max_row - 1, gfp_sheet.max_column - 1))
        for u in range(1, gfp_sheet.max_row):
            for ll in range(1, gfp_sheet.max_column):
                spts_gfp_mtx[1, u - 1, ll - 1]  =  gfp_sheet.cell(row=u + 1, column=ll + 1).value

        for tt in range(1, gfp_sheet.max_row):
            for k in range(1, gfp_sheet.max_column):
                spts_gfp_mtx[0, tt - 1, k - 1]  =  spts_sheet.cell(row=tt + 1, column=k + 3).value

        oo            =  np.where(spts_gfp_mtx[0].sum(0) == 0)[0]
        spts_gfp_mtx  =  np.delete(spts_gfp_mtx, oo, axis=2)
        tags          =  np.delete(tags, oo, axis=0)


        #### HERE WE PLOT BOTH THE VALUE, INTENSITY OF THE SPOTS AND THE GFP INTENSITY SURROUNDIN IT #####

        y_sup  =  spts_gfp_mtx.max()

        n_rows   =  6
        n_cols   =  7
        num_win  =  len(tags) // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            # time.sleep(3)
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsLayoutWidget()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(tags) - 1:
                    str_cmnd1      =  "p" + str(k) + ".plot(spts_gfp_mtx[0, :, k  + win_idxs * n_cols * n_rows], pen='r', symbol='o', symbolSize=2)"
                    str_cmnd1_bis  =  "p" + str(k) + ".plot(spts_gfp_mtx[1, :, k  + win_idxs * n_cols * n_rows], pen='b', symbol='o', symbolSize=2)"
                    str_cmnd2      =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3      =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + tags[k + win_idxs * n_cols * n_rows], color='g')"
                    str_cmnd4      =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5      =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    exec(str_cmnd1_bis)
                    exec(str_cmnd2)
                    exec(str_cmnd3)
                    exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break

        print(StrangePatch)



        #### HERE WE PLOT INTENSITY OF THE SPOTS WITH RESPECT TO THE GFP INTENSITY SURROUNDIN IT #####

        # y_sup  =  spts_gfp_mtx.max()

        n_rows   =  6
        n_cols   =  7
        num_win  =  len(tags) // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            # time.sleep(3)
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsLayoutWidget()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(tags) - 1:
                    str_cmnd1      =  "p" + str(k) + ".plot(spts_gfp_mtx[0, :, k  + win_idxs * n_cols * n_rows], spts_gfp_mtx[1, :, k  + win_idxs * n_cols * n_rows], pen='k', symbol='o', symbolSize=2)"
                    # str_cmnd1_bis  =  "p" + str(k) + ".plot(spts_gfp_mtx[1, k  + win_idxs * n_cols * n_rows, :], pen='b', symbol='o', symbolSize=2)"
                    # str_cmnd2      =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3      =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + tags[k + win_idxs * n_cols * n_rows], color='g')"
                    # str_cmnd4      =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5      =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    # exec(str_cmnd1_bis)
                    # exec(str_cmnd2)
                    exec(str_cmnd3)
                    # exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break

        print(StrangePatch)




        n_rows   =  6
        n_cols   =  7
        num_win  =  len(tags) // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            # time.sleep(3)
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsLayoutWidget()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(tags) - 1:
                    str_cmnd1      =  "p" + str(k) + ".plot(spts_gfp_mtx[1, :, k  + win_idxs * n_cols * n_rows], spts_gfp_mtx[0, :, k  + win_idxs * n_cols * n_rows], pen='k', symbol='o', symbolSize=2)"
                    # str_cmnd1_bis  =  "p" + str(k) + ".plot(spts_gfp_mtx[1, k  + win_idxs * n_cols * n_rows, :], pen='b', symbol='o', symbolSize=2)"
                    # str_cmnd2      =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3      =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + tags[k + win_idxs * n_cols * n_rows], color='g')"
                    # str_cmnd4      =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5      =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    # exec(str_cmnd1_bis)
                    # exec(str_cmnd2)
                    exec(str_cmnd3)
                    # exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break

        print(StrangePatch)





        # spts_gfp_mtx[0]  /=  spts_gfp_mtx[0].max()
        # spts_gfp_mtx[1]  /=  spts_gfp_mtx[1].max()
        #
        # corrs  =  np.zeros(spts_gfp_mtx.shape[2])
        # for yy in range(spts_gfp_mtx.shape[2]):
        #     corrs[yy]  =  np.corrcoef(spts_gfp_mtx[0, :, yy], spts_gfp_mtx[1, :, yy])[0, 1]
        #
        # spts_extr  =  np.copy(spts_gfp_mtx)
        # extr_rm    =  np.where(np.sign(spts_extr[0]).sum(0) < 25)[0]
        # spts_extr  =  np.delete(spts_extr, extr_rm, axis=2)
        #
        # corrs_ext  =  np.zeros(spts_extr.shape[2])
        # for yy in range(spts_extr.shape[2]):
        #     corrs_ext[yy]  =  np.corrcoef(spts_extr[0, :, yy], spts_extr[1, :, yy])[0, 1]

        corrs_nozer  =  []
        for hh in range(spts_gfp_mtx.shape[2]):
            a                =  spts_gfp_mtx[0, :, hh]
            b                =  spts_gfp_mtx[1, :, hh]
            a                =  a[a != 0]
            b                =  b[b != 0]
            if a.size > 15:
                corrs_nozer.append(np.corrcoef(a, b)[1, 0])
        corrs_nozer  =  np.asarray(corrs_nozer)
        print(corrs_nozer.mean())


